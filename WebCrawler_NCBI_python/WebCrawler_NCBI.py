############################################################
##  Website: https://www.ncbi.nlm.nih.gov/snp/
##  Author: BessyHuang 2022.04.12
##  Topic: Web crawler on NCBI in order to get GeneName
############################################################

from openpyxl import Workbook
from bs4 import BeautifulSoup
import requests


def request_URL(user_input):
	if ":" in user_input:
        	chrN, pos = user_input.split(":")
        	url = 'https://www.ncbi.nlm.nih.gov/snp/?term=' + chrN + '%3A' + pos
	else:
        	rsID = user_input
        	url = 'https://www.ncbi.nlm.nih.gov/snp/?term=' + rsID
	return url

def get_GeneName(soup):
	table =  soup.find('dl', attrs={'class':'snpsum_dl_left_align'})
	#print(table.string)
	try:
        	Field_Gene = soup.find(text='Gene:')
        	Value_GeneName = Field_Gene.findNext("dd").text
        
        	Value_GeneName = Value_GeneName.replace(' (Varview)\n', '')
	except:
        	Value_GeneName = "No Gene Name"

	print('==>', Field_Gene, ':', Value_GeneName)
	return Field_Gene, Value_GeneName


#1 Input keyword and send to NCBI website
user_input = input("Please input `rsID` or `chr_name:chr_position` :\n> ")
url = request_URL(user_input)
response = requests.get(url)
#print(response)


#2 Get NCBI website information
soup = BeautifulSoup(response.text, 'html.parser')
Field_Gene, Value_GeneName = get_GeneName(soup)

#3 Write to excel
wb = Workbook()
sheet1 = wb.create_sheet("GeneTable")
sheet1.cell(row=1, column=1, value=user_input)
sheet1.cell(row=3, column=1, value="GeneName")
sheet1.cell(row=3, column=2, value=Value_GeneName)
wb.save('output.xlsx')
